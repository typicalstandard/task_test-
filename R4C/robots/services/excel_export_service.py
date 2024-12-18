from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
from io import BytesIO
from robots.models import Robot
from django.db.models import Count, Q

class ExcelExportService:
    @staticmethod
    def generate_robots_report():
        wb = Workbook()
        standard_sheet = wb.active
        wb.remove(standard_sheet)

        models = Robot.objects.values_list('model', flat=True).distinct()
        one_week_ago = datetime.now() - timedelta(days=7)

        for model in models:
            ws = wb.create_sheet(title=model)
            ws.append(["Модель", "Версия", "Количество за неделю"])

            # группирует модели за последнюю неделю
            model_versions = Robot.objects.filter(model=model).values('version').annotate(
                count_week=Count('id', filter=Q(created__gte=one_week_ago)))

            for model_version in model_versions:
                version = model_version['version']
                count_week = model_version['count_week']
                ws.append([model, version, count_week])

                # форматирование столбцов в Excel
                column_index = 3
                column_letter = get_column_letter(column_index)
                ws.column_dimensions[column_letter].width = 20

                # центрирует последний столбик
                cell = ws.cell(row=ws.max_row, column=column_index)
                cell.alignment = Alignment(horizontal='center')

        # сохраняет данные в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
