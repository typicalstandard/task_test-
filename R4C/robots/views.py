from datetime import datetime, timedelta

from django.db.models import Count, Q
from django.http import  HttpResponseServerError
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .RobotForm import RobotForm

import json
from django.http import JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from .models import Robot
from io import BytesIO

@method_decorator(csrf_exempt, name='dispatch')
class CreateRobotAPIView(View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            form = RobotForm(data)
            if form.is_valid():
                form.save()
                return JsonResponse({'message': 'Robot created successfully!'}, status=201)
            else:
                return JsonResponse({'errors': form.errors}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


class ExportRobotsView(View):
        def get(self, request):
            try:
                wb = Workbook()

                # Удаление стандартного листа
                standard_sheet = wb.active
                wb.remove(standard_sheet)

                # Получение всех уникальных моделей роботов
                models = Robot.objects.values_list('model', flat=True).distinct()

                one_week_ago = datetime.now() - timedelta(days=7)

                for model in models:
                    # Создание нового листа для каждой модели робота
                    ws = wb.create_sheet(title=model)

                    ws.append(["Модель", "Версия", "Количество за неделю"])

                    # Установка ширины столбца "Количество за неделю"
                    column_index = 3
                    column_letter = get_column_letter(column_index)
                    ws.column_dimensions[column_letter].width = 20

                    # Получение всех версий для конкретной модели робота
                    model_versions = Robot.objects.filter(model=model).values('version').annotate(
                        count_week=Count('id', filter=Q(created__gte=one_week_ago)))

                    for model_version in model_versions:
                        version = model_version['version']
                        count_week = model_version['count_week']
                        ws.append([model, version, count_week])

                        # Центрирование чисел в столбце "Количество за неделю"
                        cell = ws.cell(row=ws.max_row, column=column_index)
                        cell.alignment = Alignment(horizontal='center')

                # Сохранение Excel файла в памяти
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Создание HTTP ответа с Excel файлом
                response = HttpResponse(output,
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="all_robots.xlsx"'
                print("HTTP ответ успешно создан")
                return response

            except Exception as e:
                return HttpResponseServerError("An error occurred while generating the Excel file.")
